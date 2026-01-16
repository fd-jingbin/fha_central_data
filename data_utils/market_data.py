# region Import Modules

# External Imports
import os
import subprocess
import time
import openpyxl
import numpy as np
import pandas as pd
import psutil
import win32com.client
from datetime import date
from typing import Any, List, Sequence
import shutil, tempfile

# Internal Imports
import importlib as imp
import data_config as da_cfg
import data_utils.utils as da_ut

for pkg in [da_cfg, da_ut]:
    imp.reload(pkg)

# endregion

def chunk(lst: Sequence[Any], n: int) -> List[List[Any]]:
    if n <= 0:
        raise ValueError("n must be > 0")
    return [list(lst[i: i + n]) for i in range(0, len(lst), n)]


def fast_read_excel(file_path, header=0):
    """
    并强制将空字符串 '' 转换为 np.nan，保证后续 dropna 正常工作
    """

    # wb = CalamineWorkbook.from_path(file_path)
    # rows = wb.get_sheet_by_index(0).to_python()
    #
    # if not rows:
    #     return pd.DataFrame()
    #
    # if header is None:
    #     df = pd.DataFrame(rows)
    # else:
    #     df = pd.DataFrame(rows[1:], columns=rows[0])

    if header is None:
        df = pd.read_excel(file_path, header=None)
    else:
        df = pd.read_excel(file_path)

    # 2. === 核心修复 ===
    # 这一步非常快，将所有的空字符串替换为标准的 NaN
    # 这样你后面的 pd.isna() 和 dropna() 就能生效了
    obj_cols = df.select_dtypes(include="object").columns
    if len(obj_cols) > 0:
        df[obj_cols] = df[obj_cols].replace("", np.nan)

    return df

class BbgExcelLoader:

    def __init__(self, cfg=da_cfg):
        self.cfg = cfg
        self.consensus_field_map = {
            "IS902": "Revenue",
            "IM092": "Gross Margin%",
            "F1396": "EPS Adj+",
            "IM131": "EBITDA",
            "IM132": "Operating Profit"
        }

    def export_bql_consensus_rows(
            self,
            as_of: date,
            tickers: list[str], # e.g. "6273 JP EQUITY"
            fields: list[str] = ['IS902', 'IM092', 'F1396', 'IM131', 'IM132'], # e.g. "IS902"
            value_types: list[str] = ['MAX', 'MIN', 'AVG'], # e.g. "AVG"
            periods: list[str] = ['A'], # e.g. "A"
            fprs: list[str] = ['2025Y', '2026Y', '2027Y'],  # e.g. "2026Y"
            fpo: int = 0,
            fa_act_est_data: str = "E",
            act_est_mapping: str = "PRECISE",
            fs: str = "MR",
            xlfill: str = "b",
            filename="output.xlsx"
    ) -> None:

        wb = openpyxl.Workbook()
        ws = wb.active

        # Header
        ws.append(["Date", "Ticker", "Primary", "Field", "ValueType", "Period", "FY", "Value", "Currency"])

        # Rows
        idx = 2
        for t in tickers:
            for field in fields:
                for value_type in value_types:
                    for period in periods:
                        for fpr in fprs:
                            formula_primary = f'=BDP("{t}", "EQY_FUND_TICKER") & " EQUITY"'

                            formula_value = (
                                f'=BQL(C{idx}, "CONTRIBUTOR_STATS({field}, {value_type})", '
                                f'"FPR={fpr}", "FPT={period}", "FPO={fpo}", '
                                f'"FA_ACT_EST_DATA={fa_act_est_data}", "ACT_EST_MAPPING={act_est_mapping}", '
                                f'"FS={fs}", "XLFILL={xlfill}")'
                            )

                            formula_currency = f'=BDP("{t}", "EQY_FUND_CRNCY")'
                            ws.append([as_of.isoformat(), t, formula_primary,
                                       self.consensus_field_map[field], value_type, period, fpr,
                                       formula_value, formula_currency])
                            idx += 1

        # Save the workbook
        wb.save(str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, filename)))

    def insert_names_and_formulas_simple_bdh(self, names, field, start_date, end_date, extra_param=None,
                                             filename="output.xlsx"):
        """
        Inserts names into row 1 at every second column starting from column 2.
        Inserts BDH formulas in row 2 at every second column starting from column 1.

        Parameters:
        names (list): List of bbg tickers to insert.
        field (str): FIELD parameter for the BDH formula.
        start_date (str): Start date for the BDH formula.
        end_date (str): End date for the BDH formula.
        filename (str): Name of the Excel file to save.
        """
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Insert names and formulas
        for i, name in enumerate(names):
            name_col = 2 + i * 2  # Column for the name (2, 4, 6, ...)
            ws.cell(row=1, column=name_col, value=name)

            # BDH formula inserted in the previous column (1, 3, 5, ...)
            if extra_param is not None:
                formula = f'=BDH({ws.cell(row=1, column=name_col).coordinate}, "{field}", "{start_date}", "{end_date}"'
                for pm in extra_param:
                    formula += f', "{pm}"'
                formula += ')'
            else:
                formula = f'=BDH({ws.cell(row=1, column=name_col).coordinate}, "{field}", "{start_date}", "{end_date}")'
            ws.cell(row=2, column=name_col - 1, value=formula)

        # Save the workbook
        wb.save(str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, filename)))

    def open_like_human_wait_and_save(self, input_filename, output_filename, wait_seconds=20):
        """
        Opens an Excel file like a human (simulating double-click), waits for Bloomberg refresh,
        then saves a copy with values only (removing all formulas). Always overrides the output file.
        """

        # Open Excel by simulating double-click (ensuring it opens in a separate process)
        input_fn = str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, input_filename))
        output_fn = str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, output_filename))
        subprocess.Popen(["start", "", input_fn], shell=True)

        # Wait to ensure Excel fully loads
        time.sleep(wait_seconds)

        excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # opens Excel
        wb = excel.Workbooks.Open(input_fn)
        wb.Save()
        wb.Close()
        excel.Quit()
        time.sleep(3)
        print('Excel Saved Successfully...')

    def insert_names_with_formulas_simple_bdp(self, tickers, field_dict, output_file="output.xlsx"):
        """
        Creates an Excel file with Bloomberg BDP formulas.

        Parameters:
        - tickers (list): List of Bloomberg ticker symbols.
        - field_dict (dict): Dictionary where keys are column headers (field names) and values are Bloomberg field values.
        - output_file (str): Name of the output Excel file.
        """
        # Initialize DataFrame with tickers
        df = pd.DataFrame({"Ticker": tickers})

        # Add Bloomberg formulas for each field
        for field_name, field_value in field_dict.items():
            df[field_name] = df["Ticker"].apply(
                lambda x: f'=BDP(A{df.index[df["Ticker"] == x].tolist()[0] + 2}, "{field_value}")')

        # Save to Excel
        out_dir = str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, output_file))
        with pd.ExcelWriter(out_dir, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Bloomberg Data", index=False)

        print('Excel Saved')

    def insert_names_with_formulas_simple_bds(self, names_list, field_value, extra_params, n_space,
                                              filename="output.xlsx"):
        """
        Inserts names into row 1 at every n_space column interval.
        Inserts BDS formulas below each name.

        Parameters:
        names_list (list): List of names to insert.
        field_value (str): FIELD parameter for the BDS formula.
        n_space (int): Column spacing for names.
        filename (str): Name of the Excel file to save.
        """
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active

        # Insert names and formulas
        for i, name in enumerate(names_list):
            col = 1 + i * n_space  # Column position based on n_space
            ws.cell(row=1, column=col, value=name)  # Insert name in row 1

            # Insert the BDS formula in row 2, right below the name
            formula = f'=BDS({ws.cell(row=1, column=col).coordinate}, "{field_value}", "{extra_params}")'
            ws.cell(row=2, column=col, value=formula)

        # Save the workbook
        wb.save(str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, filename)))

    def clean_horizontal_listed_bdh_data_by_ticker(self, value_type_str: str, excel_file_dir: str,
                                                   save_pickle_dir: str):
        """
        读取时自动去除每一对列末尾的 NaN 空值。
        """
        excel_file_dir = str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, excel_file_dir))
        save_pickle_dir = str(os.path.join(self.cfg.BBG_DEFAULT_PICKLE_DIR, save_pickle_dir))

        dst_dir = excel_file_dir.replace('.xlsx', '_copy.xlsx')
        tmpdir = tempfile.mkdtemp()
        shutil.copy2(excel_file_dir, dst_dir)

        try:
            df = fast_read_excel(dst_dir)
        except (TypeError, ValueError):
            # 降级处理
            df = pd.read_excel(dst_dir)

        stacked_data = []
        n_cols = df.shape[1]

        # 2. 循环遍历每一对列 (Date, Value)
        for i in range(0, n_cols, 2):
            if i + 1 >= n_cols:
                break

            # 获取列名 (Ticker Name)
            # 注意：彭博导出的 Excel，通常 Ticker 只写在 Value 那一列的表头
            ticker_col_name = df.columns[i + 1]

            # 获取原始数据 (包含末尾的 NaN)
            raw_dates = df.iloc[:, i].values
            raw_values = df.iloc[:, i + 1].values

            # === 核心修改：使用 Numpy Mask 去除空值 ===
            # 逻辑：只要日期列是空的 (NaN 或 NaT)，这一行对这个 Ticker 就是废的
            # 这会自动把 Ticker B 多出来的 4750 行空值切掉
            valid_mask = ~pd.isna(raw_dates)

            # 应用掩码，只保留有效数据
            clean_dates = raw_dates[valid_mask]
            clean_values = raw_values[valid_mask]

            # 如果这一列全是空的，直接跳过
            if len(clean_dates) == 0:
                continue

            # 3. 构建该 Ticker 专属的小 DataFrame
            temp_df = pd.DataFrame({
                'Date': clean_dates,
                'Ticker': ticker_col_name,
                value_type_str: clean_values
            })

            stacked_data.append(temp_df)

        # 4. 合并所有 Ticker 的数据
        if not stacked_data:
            return pd.DataFrame()

        df_stacked = pd.concat(stacked_data, ignore_index=True)

        # 再次确保没有脏数据 (虽然上面已经清洗过，双保险)
        df_stacked = df_stacked.dropna()

        if save_pickle_dir is not None:
            df_stacked.to_pickle(save_pickle_dir)

        return df_stacked

    def save_bdp_excel_data(self, file_path, save_pickle_dir: str = None, return_df=False):
        file_path = str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, file_path))
        save_pickle_dir = str(os.path.join(self.cfg.BBG_DEFAULT_PICKLE_DIR, save_pickle_dir))
        data = fast_read_excel(file_path)
        if save_pickle_dir is not None:
            data.to_pickle(save_pickle_dir)
        return data

    def save_bdh_to_structured_pickle(self, name_list, field_value, extra_params, start_date, end_date, temp_excel_path,
                                      final_excel_path, pickle_path, return_df=False, batch_size=150, project_path=None):

        if project_path is None:
            project_path = self.cfg.BBG_DEFAULT_PICKLE_DIR

        list_of_tickers = chunk(name_list, batch_size)
        out_list = []

        for idx, tickers in enumerate(list_of_tickers):
            print('insert formulas')
            self.insert_names_and_formulas_simple_bdh(tickers, field_value, start_date, end_date, extra_params,
                                                      filename=temp_excel_path)
            print('open and wait')
            self.open_like_human_wait_and_save(temp_excel_path, final_excel_path,
                                          wait_seconds=max(int(len(tickers) * 0.1), 30))
            print('clean and save')
            self.clean_horizontal_listed_bdh_data_by_ticker(value_type_str=field_value, excel_file_dir=temp_excel_path,
                                                       save_pickle_dir=f'{pickle_path}'.replace('.pkl',f'{idx}.pkl'))
            out_list.append(pd.read_pickle(str(os.path.join(project_path, f'{pickle_path}'.replace('.pkl',f'{idx}.pkl')))))

        out = pd.concat(out_list, ignore_index=True)
        out.to_pickle(str(os.path.join(project_path, pickle_path)))

        if return_df:
            return out

    def save_bql_consensus_data_pickle(self, name_list, temp_excel_path, final_excel_path, pickle_path, return_df=False):
        self.export_bql_consensus_rows(date.today(), name_list, filename=temp_excel_path)
        self.open_like_human_wait_and_save(temp_excel_path, final_excel_path,
                                      wait_seconds=max(int(len(name_list) * 0.1), 20))
        out = pd.read_excel(final_excel_path)
        out.to_pickle(pickle_path)
        if return_df:
            return out

    def stack_bds_excel_data(self, file_path, n_space, save_pickle_dir: str = None):
        """
        Reads an Excel file created by `insert_names_with_formulas_simple_bds` and stacks the data
        into a long-format DataFrame.

        Parameters:
        -----------
        file_path : str
            Path to the Excel file.

        Returns:
        --------
        pandas.DataFrame
            A long-format DataFrame with columns: ['Ticker', 'Field', 'Value'].

        Notes:
        ------
        - Assumes tickers are in the first row (row index 0).
        - Data under each ticker is stacked into a single DataFrame.
        - The second row is ignored since it contains Bloomberg BDS formulas.

        Example:
        --------
        Input Excel:
            | AAPL US Equity |  | MSFT US Equity |  |
            | =BDS(A1, "PX_LAST") |  | =BDS(C1, "PX_LAST") |  |
            | Revenue  | Value  | Revenue  | Value  |
            | 2022     | 300B   | 2022     | 250B   |
            | 2023     | 320B   | 2023     | 270B   |

        Output DataFrame:
            | Ticker         | Field   | Value |
            |---------------|---------|-------|
            | AAPL US Equity | Revenue | 300B  |
            | AAPL US Equity | 2022    | 300B  |
            | AAPL US Equity | 2023    | 320B  |
            | MSFT US Equity | Revenue | 250B  |
            | MSFT US Equity | 2022    | 250B  |
            | MSFT US Equity | 2023    | 270B  |
        """
        # Load the Excel file
        file_path = str(os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, file_path))
        save_pickle_dir = str(os.path.join(self.cfg.BBG_DEFAULT_PICKLE_DIR, save_pickle_dir))

        dst_dir = file_path.replace('.xlsx', '_copy.xlsx')
        tmpdir = tempfile.mkdtemp()
        shutil.copy2(file_path, dst_dir)

        df = fast_read_excel(file_path, header=None)

        # Get ticker names from row 1
        ticker_cols = df.iloc[0].dropna().index  # Get valid ticker column positions
        tickers = df.iloc[0, ticker_cols]  # Get ticker names

        stacked_data = []

        # Process each ticker's data
        for col, ticker in zip(ticker_cols, tickers):
            temp_df = df.iloc[2:, col:col + n_space]  # Skip first two rows (ticker + formula)
            temp_df.columns = df.iloc[1, col:col + n_space]  # Rename columns
            temp_df['Ticker'] = ticker  # Add ticker column
            stacked_data.append(temp_df)

        # Concatenate all ticker data
        df_stacked = pd.concat(stacked_data, ignore_index=True)

        if 'Announcement Date' in df_stacked:
            df_stacked = df_stacked.dropna(subset='Announcement Date')

        if save_pickle_dir is not None:
            df_stacked.to_pickle(save_pickle_dir)

        return df_stacke.dropna(axis=1, how="all")

    def save_bds_to_structured_pickle(self, name_list, field_value, extra_params, n_space, temp_excel_path, final_excel_path,
                                      pickle_path, return_df=False):
        self.insert_names_with_formulas_simple_bds(name_list, field_value, extra_params, n_space, filename=temp_excel_path)
        self.open_like_human_wait_and_save(temp_excel_path, final_excel_path,
                                      wait_seconds=max(int(len(name_list) * 0.5), 20))
        self.stack_bds_excel_data(file_path=temp_excel_path, n_space=n_space, save_pickle_dir=pickle_path)
        if return_df:
            return pd.read_pickle(str(os.path.join(self.cfg.BBG_DEFAULT_PICKLE_DIR, pickle_path)))

    def save_bdp_to_structured_pickle(self, name_list, field_dict, temp_excel_path, final_excel_path, pickle_path, project_path=None, return_df=False):

        self.insert_names_with_formulas_simple_bdp(name_list, field_dict, output_file=temp_excel_path)
        self.open_like_human_wait_and_save(temp_excel_path, final_excel_path,
                                      wait_seconds=max(int(len(name_list) * 0.02), 20))
        out = self.save_bdp_excel_data(file_path=temp_excel_path, save_pickle_dir=pickle_path)
        if return_df:
            return out

    def load_saved_bbg_pickle(self, pkl_name):
        return pd.read_pickle(str(os.path.join(self.cfg.BBG_DEFAULT_PICKLE_DIR, f'{pkl_name}.pkl')))

    def get_ohlc_data(self, ticker_list, extra_params, start_date, end_date, extra_fields=None):
        df_list = []
        if extra_fields is not None and isinstance(extra_fields, list):
            field_list = ['PX_OPEN', 'PX_HIGH', 'PX_LOW', 'PX_LAST'] + extra_fields
        else:
            field_list = ['PX_OPEN', 'PX_HIGH', 'PX_LOW', 'PX_LAST']
        for field_val in field_list:
            df = self.save_bdh_to_structured_pickle(
                ticker_list, field_value=field_val, start_date=start_date, end_date=end_date, extra_params=extra_params,
                temp_excel_path=os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, f'TEMP_{field_val}.xlsx'),
                final_excel_path=os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, f'FINAL_{field_val}.xlsx'),
                pickle_path=os.path.join(self.cfg.BBG_EXCEL_TEMP_FILE_DIR, f'TEMP_{field_val}.pkl'),
                return_df=True)
            df_list.append(df)
        return pd.concat([d.set_index(['Date', 'Ticker']) for d in df_list], axis=1)