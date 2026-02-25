from pathlib import Path
import re
import os
import time
import uuid
import glob
import pandas as pd
from datetime import datetime
from typing import Dict, Tuple, Union, Any, Optional, List
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
import win32com.client as win32
import json
import re



_CellPos = Tuple[pd.DataFrame, int, Union[int, str], bool, bool]
SheetMap = Dict[str, Dict[str, _CellPos]]


def _parse_start(start_row: int, start_col: Union[int, str]) -> Tuple[int, int]:
    """
    Convert user input to 1-based Excel (row, col).
    Accepts either (row, col_int) or (ignored_row, "B4").
    """
    if isinstance(start_col, str):                                          # "B4"
        m = re.match(r"^([A-Z]+)(\d+)$", start_col)
        if not m:
            raise ValueError(f"Bad cell reference: {start_col!r}")
        col0 = column_index_from_string(m.group(1))
        row0 = int(m.group(2))
    else:                                                                   # numbers
        col0 = int(start_col)
        row0 = int(start_row)
    return row0, col0


def _set_value_allow_merged(ws: Worksheet, row: int, col: int, value) -> None:
    """
    Safely set cell value even if target is inside a merged range.
    If (row, col) is merged, redirect to that range's top-left anchor.
    """
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    # find merged range containing the coordinate
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            ws.cell(row=rng.min_row, column=rng.min_col, value=value)
            return

    # fall-back (should not happen)
    cell.value = value


def _flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with flattened column names if MultiIndex present."""
    if isinstance(df.columns, pd.MultiIndex):
        df = df.copy()
        df.columns = [' / '.join(map(str, col)).strip() for col in df.columns]
    return df
# ------------------------------------------------------------------ #


def _dump_dataframe(ws: Worksheet,
                    df: pd.DataFrame,
                    row0: int, col0: int,
                    write_header: bool, write_index: bool) -> None:
    """
    Write *one* DataFrame into an openpyxl Worksheet starting at (row0, col0).
    """
    df = _flatten_columns(df)

    data_row0 = row0 + (1 if write_header else 0)

    # -------- header --------
    if write_header:
        c = col0
        if write_index:
            idx_name = df.index.name or ""
            _set_value_allow_merged(ws, row0, c, idx_name)
            c += 1
        for col_name in df.columns:
            _set_value_allow_merged(ws, row0, c, col_name)
            c += 1

    # -------- data rows -----
    for i, row in enumerate(df.itertuples(index=False), start=0):
        r = data_row0 + i
        c = col0
        if write_index:
            ws.cell(row=r, column=c, value=df.index[i])          # data area is rarely merged
            c += 1
        for v in row:
            ws.cell(row=r, column=c, value=v)
            c += 1


def write_dfs_to_template(template_path: str,
                          output_path: str,
                          sheet_map: SheetMap) -> None:
    """
    Parameters
    ----------
    template_path : str
        Path to an existing .xlsx template.
    output_path : str
        Where to save the filled workbook.
    sheet_map : dict
        {sheet_name: {identifier: (df, start_row, start_col, write_header, write_index), …}, …}

        `start_col` may be an int **or** Excel address like "B4".
    """
    wb = load_workbook(template_path)

    for sheet, tbl_map in sheet_map.items():
        if sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in template.")
        ws = wb[sheet]

        for name, (df, start_row, start_col,
                   write_header, write_index) in tbl_map.items():

            row0, col0 = _parse_start(start_row, start_col)
            _dump_dataframe(ws, df, row0, col0, write_header, write_index)

            print(f"[{sheet}] wrote '{name}' "
                  f"{df.shape[0]}×{df.shape[1]} at ({row0},{col0}), "
                  f"hdr={write_header}, idx={write_index}")

    wb.save(output_path)
    print(f"Saved workbook ➜ {output_path}")


def inner_level_as_row(df, inner_level: int = 1, row_name: str = None, top=True):
    """
    Keep outer-level columns; turn the chosen *inner_level* into a single data row.

    Parameters
    ----------
    df : pandas.DataFrame               # must have at least 2 column levels
    inner_level : int | str, default 1  # which level to pull down
    row_name : str | None               # index label for the new row;
                                         # default = df.columns.names[inner_level] or 'inner'
    top : bool, default True            # place the new row at the beginning (top) or end

    Returns
    -------
    DataFrame with:
      • columns  = outer level values             (MultiIndex collapsed)
      • new row  = inner level values
      • all original rows preserved unchanged
    """
    if not isinstance(df.columns, pd.MultiIndex) or df.columns.nlevels < 2:
        raise ValueError("Input must have ≥2 column levels")

    # Pull outer & inner levels once so order is preserved
    outer_vals = df.columns.get_level_values(0)
    inner_vals = df.columns.get_level_values(inner_level)

    # Decide the new row's index label
    if row_name is None:
        names = df.columns.names
        row_name = names[inner_level] if names and names[inner_level] else "inner"

    # 1️⃣  collapse columns to outer level
    out = df.copy()
    out.columns = outer_vals

    # 2️⃣  build the Series that becomes the new row
    new_row = pd.Series(inner_vals, index=out.columns, name=row_name)

    # 3️⃣  concat on top or bottom
    if top:
        out = pd.concat([new_row.to_frame().T, out])
    else:
        out.loc[row_name] = new_row

    return out


def exist_file(file_path):
    return Path(file_path).exists()


def join_path_name(path, name):
    return os.path.join(path, name)


def exist_create_folder(folder_path):
    Path(folder_path).mkdir(parents=True, exist_ok=True)


def find_avail_files(folder_path, keywords, file_format):
    avail_files = glob.glob(os.path.join(folder_path, f"{keywords}*.{file_format}"))
    if not avail_files:
        raise Exception("No matching files found.")
    return avail_files


def find_avail_folders(folder_path, keywords=""):
    pattern = os.path.join(folder_path, f"{keywords}*")
    matches = glob.glob(pattern)

    avail_folders = [p for p in matches if os.path.isdir(p)]
    if not avail_folders:
        raise Exception("No matching folders found.")
    return avail_folders


def find_avail_files_sparse_search(folder_path, keywords, file_format):
    file_format = file_format.lower().lstrip(".")
    keywords_lower = keywords.lower()

    matched_files = []
    for root, _, files in os.walk(folder_path):
        for file in files:
            # check extension and keyword
            if file.lower().endswith(f".{file_format}") and any(kw in file.lower() for kw in keywords_lower):
                matched_files.append(os.path.join(root, file))

    return matched_files


def find_latest_file(folder_path, keywords, file_format):
    latest_file = max(find_avail_files(folder_path, keywords, file_format), key=os.path.getmtime)
    return latest_file


def find_earliest_enfusion_daily(folder_path, target_date, latest=False):
    pattern = re.compile(r'_(\d{2})_(\d{2})_(\d{4})_(\d{2})_(\d{2})_(\d{2})\.xls$', re.IGNORECASE)
    target_dt = datetime.strptime(target_date, '%Y-%m-%d')

    matched_files = []

    for filename in os.listdir(folder_path):
        match = pattern.search(filename)
        if match:
            month, day, year, hour, minute, second = map(int, match.groups())
            file_dt = datetime(year, month, day, hour, minute, second)

            if file_dt.date() == target_dt.date():
                matched_files.append((file_dt, filename))

    if not matched_files:
        print(f"No files found for {target_date} in {folder_path}.")
        return None

    matched_files.sort()

    if latest:
        return matched_files[-1][1]
    else:
        return matched_files[0][1]


def insert_additional_second_level_columns(df1, df2, new_col_nm):
    new_col_set = pd.MultiIndex.from_product([df2.columns, [new_col_nm]], names=df1.columns.names)
    df2 = df2.reindex(df1.index)
    df2.columns = new_col_set
    df1 = pd.concat([df1, df2], axis=1).sort_index(axis=1, level=[0, 1]) \
    .pipe(lambda d: d[[(o, i) for o in pd.Index(pd.unique(df1.columns.get_level_values(0))).tolist() for i in df1[df1.columns.levels[0][0]].columns.tolist() + [new_col_nm] if (o, i) in d.columns]])
    return df1


def excel_to_pdf_landscape_fitted(excel_path: str, output_pdf: str, auto_fit_sheets: list = None, portrait_sheets: list = None):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(excel_path)
    auto_fit_sheets = [] if auto_fit_sheets is None else auto_fit_sheets
    portrait_sheets = [] if portrait_sheets is None else portrait_sheets
    for sheet in wb.Sheets:
        print(sheet.Name)
        if sheet.Name.lower() in portrait_sheets:
            sheet.PageSetup.Orientation = 1
        else:
            sheet.PageSetup.Orientation = 2  # 2 = Landscape
        sheet.PageSetup.Zoom = False     # disable Zoom so FitToPages works
        sheet.PageSetup.FitToPagesWide = 1

        if sheet.Name.lower() in auto_fit_sheets:
            sheet.PageSetup.FitToPagesTall = False  # 0/False ⇒ automatic height
        else:
            sheet.PageSetup.FitToPagesTall = 1

    # Export entire workbook as PDF
    wb.ExportAsFixedFormat(0, output_pdf)  # 0 = PDF format

    wb.Close(False)
    excel.Quit()
    print(f"Saved PDF: {output_pdf}")


def get_latest_file_name(search_dir, search_name):
    csv_files = glob.glob(os.path.join(search_dir, search_name))
    if not csv_files:
        raise Exception("No CSV files found.")
    latest_file = max(csv_files, key=os.path.getmtime)
    return latest_file


def check_file_modified_date(fl_dir):
    """
    Returns the last modified date of a file.

    Parameters
    ----------
    fl_dir : str
        Path to the file.

    Returns
    -------
    datetime
        Last modified datetime of the file.
    """
    if not os.path.isfile(fl_dir):
        raise FileNotFoundError(f"File not found: {fl_dir}")

    modified_timestamp = os.path.getmtime(fl_dir)
    return datetime.fromtimestamp(modified_timestamp)


def read_json(p):
    p = Path(p)
    try:
        return json.loads(p.read_text(encoding='utf-8'))
    except FileNotFoundError:
        raise FileNotFoundError(f"JSON file not found: {p}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in {p}: {e}")


def write_excel_atomic(df, final_path, **to_excel_kwargs):
    tmp_path = f"{final_path}.{uuid.uuid4().hex}.tmp.xlsx"
    # 1) 写到临时文件，并确保关闭
    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, **to_excel_kwargs)
        # with 退出会自动 writer.close()

    # 2) 原子替换：Windows 上也可以用 os.
    os.replace(tmp_path, final_path)



def extract_text_only(obj: Any) -> str:
    """
    从常见对象里取出 text：
    - 你的 AskResult: .text
    - OpenAI Responses对象: .output_text
    - dict: 尝试 obj["text"] / obj["output_text"]
    - 直接是 str: 原样返回
    """
    if obj is None:
        return ""

    if isinstance(obj, str):
        return obj

    # AskResult-like
    if hasattr(obj, "text"):
        t = getattr(obj, "text", "")
        return t if isinstance(t, str) else str(t)

    # OpenAI Responses-like
    if hasattr(obj, "output_text"):
        t = getattr(obj, "output_text", "")
        return t if isinstance(t, str) else str(t)

    # dict fallback
    if isinstance(obj, dict):
        for k in ("text", "output_text"):
            if k in obj:
                t = obj.get(k, "")
                return t if isinstance(t, str) else str(t)

    return str(obj)


# ---------- 2) 从 text 中解析出 “第一段完整 JSON” ----------

def _fix_duplicate_key_label(raw: str) -> str:
    # 把 "D3": "D3": "...." 修成 "D3": "...."
    return re.sub(r'("([^"]+)"\s*:\s*)"(\2)"\s*:\s*', r"\1", raw)


def _escape_inner_quotes_heuristic(s: str) -> str:
    """
    尝试修复字符串值里未转义的引号（启发式）。
    风险：极端情况可能过修；所以只作为 json.loads 失败后的补救。
    """
    out = []
    i = 0
    in_str = False
    escaped = False

    while i < len(s):
        ch = s[i]
        if not in_str:
            out.append(ch)
            if ch == '"':
                in_str = True
                escaped = False
            i += 1
            continue

        # in_str
        if escaped:
            out.append(ch)
            escaped = False
            i += 1
            continue

        if ch == "\\":
            out.append(ch)
            escaped = True
            i += 1
            continue

        if ch == '"':
            # lookahead next non-ws
            j = i + 1
            while j < len(s) and s[j] in " \t\r\n":
                j += 1
            # 如果后面是 JSON 合法分隔符，认为是字符串结束；否则当作字符串内部引号，转义
            if j >= len(s) or s[j] in [",", ":", "}", "]"]:
                out.append(ch)
                in_str = False
            else:
                out.append('\\"')
            i += 1
            continue

        out.append(ch)
        i += 1

    return "".join(out)


def _extract_fenced_block(s: str) -> str:
    """
    如果存在 ```json ... ``` 或 ``` ... ```，取“最长的 fenced block”。
    """
    s = (s or "").strip()
    blocks = re.findall(r"```(?:\s*json)?\s*\n?(.*?)\n?```", s, flags=re.I | re.S)
    if not blocks:
        return s
    return max(blocks, key=len).strip()


def _extract_first_balanced_json_chunk(s: str) -> str:
    """
    从字符串里找到第一段完整的 JSON object/array（平衡括号法）。
    """
    s = (s or "").strip()
    if not s:
        raise ValueError("Empty string")

    # 找第一个 { 或 [
    i_obj = s.find("{")
    i_arr = s.find("[")
    starts = [(i_obj, "{"), (i_arr, "[")]
    starts = [(i, ch) for i, ch in starts if i != -1]
    if not starts:
        raise ValueError("No JSON start token '{' or '[' found")

    start, opener = min(starts, key=lambda x: x[0])
    closer = "}" if opener == "{" else "]"

    depth = 0
    in_str = False
    esc = False

    for i in range(start, len(s)):
        c = s[i]

        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
            continue

        if c == '"':
            in_str = True
            continue

        if c == opener:
            depth += 1
        elif c == closer:
            depth -= 1
            if depth == 0:
                return s[start : i + 1]

    raise ValueError("No complete JSON found (unbalanced braces)")


def parse_first_json_from_text(text: str) -> Union[Dict[str, Any], list]:
    """
    从 text 里解析“第一段 JSON”：
    - 支持 fenced code block
    - 支持 text 前后夹杂非 JSON
    - json.loads 失败时做温和修复：trailing comma / 奇怪引号 / duplicate key label
    """
    raw = _extract_fenced_block(text)

    # 去掉外层引号（有些模型会把 JSON 当字符串返回）
    raw_strip = raw.strip()
    if len(raw_strip) >= 2 and raw_strip[0] == raw_strip[-1] and raw_strip[0] in ("'", '"'):
        raw_strip = raw_strip[1:-1].strip()

    # 先尝试直接 loads
    try:
        return json.loads(raw_strip)
    except json.JSONDecodeError:
        pass

    # 抽取第一段完整 JSON chunk
    chunk = _extract_first_balanced_json_chunk(raw_strip)

    # 尝试基础清洗再 loads
    cleaned = chunk
    cleaned = cleaned.replace("\\'", "'")
    cleaned = re.sub(r",(\s*[}\]])", r"\1", cleaned)          # trailing comma
    cleaned = re.sub(r'\\\s+([\\/"bfnrtu])', r"\\\1", cleaned) # 修复类似 "\ n"
    cleaned = _escape_inner_quotes_heuristic(cleaned)
    cleaned = _fix_duplicate_key_label(cleaned)

    return json.loads(cleaned)


# ---------- 3) JSON -> DataFrame ----------

def json_text_to_dataframe(
    text: str,
    *,
    key_name: str = "key",
) -> pd.DataFrame:
    """
    把 text 中的 JSON 解析出来并转 DataFrame：
    - list -> DataFrame(list)
    - dict:
        * 如果 value 全是 dict/list/str-json：按 index 展开（行=key）
        * 否则当成单行 record
    """
    data = parse_first_json_from_text(text)

    if isinstance(data, list):
        return pd.DataFrame(data)

    if isinstance(data, dict):
        # dict-of-things：尽量按 key 做行索引
        if data and all(isinstance(v, (dict, list, str)) for v in data.values()):
            rows: Dict[str, Any] = {}
            for k, v in data.items():
                if isinstance(v, str):
                    # 尝试把字符串里的 JSON 再解析一层
                    try:
                        rows[k] = parse_first_json_from_text(v)
                    except Exception:
                        rows[k] = {"_raw": v}
                else:
                    rows[k] = v

            df = pd.DataFrame.from_dict(rows, orient="index").rename_axis(key_name).reset_index()
            return df

        # 普通 dict：单行
        return pd.DataFrame([data])

    # 兜底
    return pd.DataFrame([{"value": data}])

def zip_excel_files(src_dir: str, zip_path: str, recursive: bool = True, excel_suffix: List[str] = None) -> None:

    import zipfile

    src = Path(src_dir).resolve()
    out = Path(zip_path).resolve()

    files = src.rglob("*") if recursive else src.glob("*")
    excel_files = [
        p for p in files
        if p.is_file()
        and p.suffix.lower() in excel_suffix
        and not p.name.startswith("~$")  # skip temporary Office lock files
    ]

    out.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for f in excel_files:
            z.write(f, arcname=f.relative_to(src))  # keep folder structure

    print(f"Zipped {len(excel_files)} Excel file(s) into: {out}")